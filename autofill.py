import openpyxl

def initialize_worksheet(wb, sheet_name):

    ws = wb.active
    if ws.title == "Sheet":
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # Set the header row
    ws['B2'] = "Generation / Experiments"
    for i in range(10):
        ws.cell(row=2, column=3+i, value=i+1)

    # Fill B3:B2002 with 1 to 2000
    for i in range(2000):
        ws.cell(row=3+i, column=2, value=i+1)

    # Fill other specified cells
    ws['B2003'] = "Min"
    ws['B2004'] = "Max"
    ws['B2006'] = "Values"
    for i in range(30):
        ws.cell(row=2007+i, column=2, value=i+1)
    ws['B2038'] = "Time"
    
    return ws

def process_file(ws, file_name, column):

    try:
        file = open(file_name, 'r')
    except:
        print(f"Somethings when wrong when we are trying to open {file_name}," \
              "please check if the file is located in the correct directory")

    current_row = 3

    # Read fitness values
    for line in file:
        line = line.strip()
        if not line:
            break
        ws.cell(row=current_row, column=column, value=float(line))
        current_row += 1

    # Write MIN and MAX formulas
    ws.cell(row=2003, column=column, value=f"=MIN({chr(64+column)}3:{chr(64+column)}2002)")
    ws.cell(row=2004, column=column, value=f"=MAX({chr(64+column)}3:{chr(64+column)}2002)")

    # Skip blank lines to reach chromosome bits
    while True:
        line = file.readline().strip()
        if line:
            break

    # Read chromosome bits
    current_row = 2007
    while True:
        ws.cell(row=current_row, column=column, value=float(line))
        current_row += 1
        line = file.readline().strip()
        if not line:
            break

    # Skip blank lines to reach the running time
    while True:
        line = file.readline().strip()
        if line:
            ws.cell(row=2038, column=column, value=float(line))
            break

    file.close()

def main(workbook_filename, sheet_info):

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    for sheet_name, text_file_prefix, num_files in sheet_info:
        ws = initialize_worksheet(wb, sheet_name)
        for i in range(1, num_files + 1):
            file_name = f"{text_file_prefix}_{i}.txt"
            column = 3 + (i - 1)
            process_file(ws, file_name, column)

    # Save the workbook
    wb.save(workbook_filename)

workbook_filename = "GA_Results.xlsx"
sheet_info = [
    ("GA01-Exponential", "Exponential", 10),
    # Add more sheets if needed
    # ("Sheet2", "Prefix2", num_files2),
]

if __name__ == "__main__":
    main(workbook_filename, sheet_info)